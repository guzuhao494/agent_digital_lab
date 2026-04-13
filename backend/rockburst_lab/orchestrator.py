from __future__ import annotations

import csv
import json
import math
import statistics
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from .openclaw_adapter import AgentInvocation, OpenClawRuntime


def parse_float(value: Any, default: float = 0.0) -> float:
    try:
        return default if value is None or value == "" else float(value)
    except (TypeError, ValueError):
        return default


def parse_time(value: str) -> datetime:
    return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)


def clamp(value: float, lower: float = 0.0, upper: float = 1.0) -> float:
    return max(lower, min(upper, value))


def normalize(value: float, low: float, high: float) -> float:
    return 0.0 if high <= low else clamp((value - low) / (high - low))


def mean(values: Iterable[float]) -> float:
    data = list(values)
    return statistics.fmean(data) if data else 0.0


def stdev(values: Iterable[float]) -> float:
    data = list(values)
    return statistics.pstdev(data) if len(data) > 1 else 0.0


def safe_ratio(part: float, total: float) -> float:
    return 0.0 if total <= 0 else part / total


def relative_range(values: Iterable[float]) -> float:
    data = list(values)
    average = abs(mean(data))
    if not data or average <= 1e-9:
        return 0.0
    return (max(data) - min(data)) / average


def z_score(latest: float, values: Iterable[float]) -> float:
    data = list(values)
    sigma = stdev(data)
    if sigma <= 1e-9:
        return 0.0
    return abs(latest - mean(data)) / sigma


def risk_level(score: float) -> str:
    if score >= 0.78:
        return "严重"
    if score >= 0.58:
        return "高"
    if score >= 0.34:
        return "关注"
    return "低"


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def read_table(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        except StopIteration:
            return []

        records = []
        for row in rows:
            if row is None or not any(value is not None and value != "" for value in row):
                continue
            records.append(
                {
                    headers[index]: value
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index]
                }
            )
        workbook.close()
        return records
    raise ValueError(f"不支持的表格格式：{path.suffix}，请使用 .csv 或 .xlsx")


@dataclass
class DomainAgent:
    name: str
    role: str
    runtime: OpenClawRuntime

    def invoke(self, state: dict[str, Any], logic: Callable[[dict[str, Any]], dict[str, Any]]) -> dict[str, Any]:
        return self.runtime.invoke(
            AgentInvocation(
                name=self.name,
                role=self.role,
                prompt=f"Analyze rockburst risk as {self.role}.",
                observation=state,
            ),
            lambda: logic(state),
        )


class MicroseismicSensingAgent(DomainAgent):
    def analyze(self, state: dict[str, Any]) -> dict[str, Any]:
        return self.invoke(state, self._logic)

    def _logic(self, state: dict[str, Any]) -> dict[str, Any]:
        vector = state["state_vector"]
        features = state["state_snapshot"]["microseismic_features"]
        centroid = features["spatial_centroid"]
        high_energy_events = sorted(
            state["microseismic"]["events"],
            key=lambda event: event["energy"],
            reverse=True,
        )[:3]
        score = clamp(
            0.36 * vector["microseismic_energy_index"]
            + 0.26 * vector["microseismic_cluster_index"]
            + 0.22 * vector["microseismic_energy_acceleration"]
            + 0.16 * vector["microseismic_mechanism_index"]
        )
        findings = [
            f"近窗事件数 {features['event_count']}，累计能量 {features['cumulative_energy_j']:.2e} J",
            f"空间重心 ({centroid['x']:.1f}, {centroid['y']:.1f}, {centroid['z']:.1f})，最大能量 {features['max_energy_j']:.2e} J",
            f"掌子面前方比例 {features['front_event_ratio']:.2f}，剪切/拉伸占比 {features['mechanism_ratios']['shear']:.2f}/{features['mechanism_ratios']['tensile']:.2f}",
        ]
        if vector["microseismic_mechanism_index"] > 0.55:
            findings.append("剪切和张拉破裂信号并存，提示局部应力重分配增强")
        return {
            "score": round(score, 3),
            "level": risk_level(score),
            "microseismic_activity": round(clamp(0.42 * features["event_density_index"] + 0.38 * features["energy_index"] + 0.20 * features["energy_acceleration_index"]), 3),
            "cluster_center": centroid,
            "high_energy_events": [
                {
                    "timestamp": event["timestamp"],
                    "x": event["x"],
                    "y": event["y"],
                    "z": event["z"],
                    "energy_j": event["energy"],
                    "mechanism": event["mechanism"],
                }
                for event in high_energy_events
            ],
            "mechanism_statistics": features["mechanism_ratios"],
            "preliminary_risk_score": round(score, 3),
            "findings": findings,
            "signals": {
                "energy_index": vector["microseismic_energy_index"],
                "cluster_index": vector["microseismic_cluster_index"],
                "mechanism_index": vector["microseismic_mechanism_index"],
            },
        }


class TbmConditionAgent(DomainAgent):
    def analyze(self, state: dict[str, Any]) -> dict[str, Any]:
        return self.invoke(state, self._logic)

    def _logic(self, state: dict[str, Any]) -> dict[str, Any]:
        vector = state["state_vector"]
        features = state["state_snapshot"]["tbm_features"]
        disturbance = vector["tbm_disturbance_index"]
        condition_label = "强扰动掘进" if disturbance >= 0.58 else "异常波动掘进" if features["anomaly_score"] >= 0.42 else "平稳掘进"
        score = clamp(
            0.38 * vector["tbm_disturbance_index"]
            + 0.24 * vector["tbm_thrust_anomaly"]
            + 0.24 * vector["tbm_torque_anomaly"]
            + 0.14 * vector["tbm_advance_pressure_index"]
        )
        return {
            "score": round(score, 3),
            "level": risk_level(score),
            "condition_label": condition_label,
            "disturbance_intensity": round(disturbance, 3),
            "abnormal_excavation_fluctuation": features["fluctuation_amplitude"],
            "coupling_hint": (
                "TBM 推力/扭矩波动与微震能量抬升同步，需要降低扰动并复核前方结构"
                if disturbance >= 0.48 and state["state_snapshot"]["microseismic_features"]["energy_acceleration_index"] >= 0.4
                else "当前 TBM 扰动可控，继续观察微震能量斜率"
            ),
            "findings": [
                f"当前桩号 K{state['chainage']:.1f}，推进速度 {state['tbm']['latest']['advance_rate']:.2f} mm/min",
                f"推力异常 {vector['tbm_thrust_anomaly']:.2f}，扭矩异常 {vector['tbm_torque_anomaly']:.2f}，扰动指数 {vector['tbm_disturbance_index']:.2f}",
            ],
            "signals": {
                "disturbance_index": vector["tbm_disturbance_index"],
                "advance_pressure_index": vector["tbm_advance_pressure_index"],
            },
        }


class GeologyCognitionAgent(DomainAgent):
    def analyze(self, state: dict[str, Any]) -> dict[str, Any]:
        return self.invoke(state, self._logic)

    def _logic(self, state: dict[str, Any]) -> dict[str, Any]:
        vector = state["state_vector"]
        features = state["state_snapshot"]["geology_features"]
        active = state["geology"]["active_structures"]
        score = clamp(
            0.42 * vector["geology_structure_index"]
            + 0.34 * vector["geology_stress_index"]
            + 0.24 * vector["geology_brittleness_index"]
        )
        sources = [f"{item['type']} K{item['chainage_start']:.0f}-K{item['chainage_end']:.0f}" for item in active]
        if not sources:
            sources = ["当前窗未命中显著构造，但仍保留高地应力背景权重"]
        return {
            "score": round(score, 3),
            "level": risk_level(score),
            "current_geology_summary": features["current_geologic_body_type"],
            "structural_risk_tags": features["structural_risk_tags"],
            "risk_modifier": round(0.85 + 0.42 * features["geologic_complexity_score"], 3),
            "findings": [
                f"构造风险源：{'；'.join(sources)}",
                f"最大主应力 {state['geology']['in_situ_stress_mpa']:.1f} MPa，岩性脆性指数 {vector['geology_brittleness_index']:.2f}",
            ],
            "signals": {
                "structure_index": vector["geology_structure_index"],
                "stress_index": vector["geology_stress_index"],
            },
        }


class MechanismMatchingAgent(DomainAgent):
    def analyze(self, state: dict[str, Any]) -> dict[str, Any]:
        return self.invoke(state, self._logic)

    def _logic(self, state: dict[str, Any]) -> dict[str, Any]:
        vector = state["state_vector"]
        mechanisms = [
            {
                "name": "高地应力-脆性破裂型",
                "path": "高地应力蓄能 -> 脆性围岩损伤 -> 微裂纹扩展 -> 能量释放",
                "score": 0.36 * vector["geology_stress_index"]
                + 0.28 * vector["geology_brittleness_index"]
                + 0.22 * vector["microseismic_energy_index"]
                + 0.14 * vector["microseismic_mechanism_index"],
                "evidence": [
                    f"最大主应力指数 {vector['geology_stress_index']:.2f}",
                    f"脆性指数 {vector['geology_brittleness_index']:.2f}",
                    f"微震能量指数 {vector['microseismic_energy_index']:.2f}",
                ],
            },
            {
                "name": "结构面滑移扰动型",
                "path": "结构面切割 -> TBM 扰动加载 -> 剪切滑移 -> 局部块体失稳",
                "score": 0.40 * vector["geology_structure_index"]
                + 0.24 * vector["tbm_disturbance_index"]
                + 0.22 * vector["microseismic_cluster_index"]
                + 0.14 * vector["microseismic_mechanism_index"],
                "evidence": [
                    f"构造复杂度 {vector['geology_structure_index']:.2f}",
                    f"TBM 扰动指数 {vector['tbm_disturbance_index']:.2f}",
                    f"空间聚集指数 {vector['microseismic_cluster_index']:.2f}",
                ],
            },
            {
                "name": "掘进卸荷-能量突释型",
                "path": "快速掘进卸荷 -> 应力重分布 -> 微震能量加速 -> 动态释放",
                "score": 0.34 * vector["tbm_advance_pressure_index"]
                + 0.30 * vector["microseismic_energy_acceleration"]
                + 0.22 * vector["tbm_torque_anomaly"]
                + 0.14 * vector["microseismic_energy_index"],
                "evidence": [
                    f"推进压力指数 {vector['tbm_advance_pressure_index']:.2f}",
                    f"能量加速指数 {vector['microseismic_energy_acceleration']:.2f}",
                    f"扭矩异常 {vector['tbm_torque_anomaly']:.2f}",
                ],
            },
        ]
        scored = [
            {
                **item,
                "score": round(clamp(item["score"]), 3),
                "level": risk_level(item["score"]),
            }
            for item in mechanisms
        ]
        dominant = max(scored, key=lambda item: item["score"])
        path = dominant["path"]
        score = dominant["score"]
        nodes = state["geology"].get("mechanism_graph", {}).get("nodes", [])
        matched_nodes = [node["id"] for node in nodes if node.get("trigger_threshold", 0) <= score + 0.1][:4]
        return {
            "score": round(clamp(score), 3),
            "level": risk_level(score),
            "candidate_mechanisms": scored,
            "mechanism_scores": {item["name"]: item["score"] for item in scored},
            "dominant_mechanism": dominant["name"],
            "dominant_path": path,
            "trigger_evidence": dominant["evidence"],
            "matched_nodes": matched_nodes,
            "findings": [
                f"主导机理：{dominant['name']}",
                f"主导机理路径：{path}",
                f"机理匹配强度 {score:.2f}，匹配节点 {', '.join(matched_nodes) if matched_nodes else '待补充'}",
            ],
        }


class SimulationExperimentAgent(DomainAgent):
    def analyze(self, state: dict[str, Any]) -> dict[str, Any]:
        return self.invoke(state, self._logic)

    def _logic(self, state: dict[str, Any]) -> dict[str, Any]:
        vector = state["state_vector"]
        chainage = state["chainage"]
        base_score = clamp(
            0.30 * vector["microseismic_energy_index"]
            + 0.22 * vector["microseismic_cluster_index"]
            + 0.20 * vector["tbm_disturbance_index"]
            + 0.18 * vector["geology_structure_index"]
            + 0.10 * vector["geology_stress_index"]
        )
        scenarios = [
            {
                "scenario_key": "scenario_a",
                "name": "场景 A：维持当前掘进参数",
                "factor": 1.00,
                "bias": 0.02,
                "suggested_action": "维持当前推进参数，仅提高微震监测刷新频率",
            },
            {
                "scenario_key": "scenario_b",
                "name": "场景 B：降低推进速度 10%",
                "factor": 0.86,
                "bias": -0.02,
                "suggested_action": "推进速度下调 10%，同步监控推力和微震能量斜率",
            },
            {
                "scenario_key": "scenario_c",
                "name": "场景 C：降低推进速度并加强局部支护",
                "factor": 0.72,
                "bias": -0.04,
                "suggested_action": "推进速度下调 10%，支护提前量和局部支护强度同步提高",
            },
            {
                "scenario_key": "scenario_d",
                "name": "场景 D：前方 10m 存在不利结构面",
                "factor": 1.12,
                "bias": 0.07,
                "suggested_action": "立即开展超前探测，并准备短进尺弱扰动通过",
            },
            {
                "scenario_key": "scenario_e",
                "name": "场景 E：地质扰动减弱",
                "factor": 0.80,
                "bias": -0.03,
                "suggested_action": "保持保守推进，若监测验证扰动减弱则逐步恢复参数",
            },
        ]
        branches = []
        acceleration = vector["microseismic_energy_acceleration"]
        structure_memory = 0.05 * vector["geology_structure_index"]
        for scenario in scenarios:
            curve = []
            risk_scores = {}
            for hours, drift_weight in ((1, 0.055), (3, 0.115), (6, 0.18)):
                drift = drift_weight * acceleration + structure_memory
                score = clamp(base_score * scenario["factor"] + scenario["bias"] + drift)
                label = f"{hours}h"
                risk_scores[label] = round(score, 3)
                curve.append({"window": f"T+{label}", "risk_score": round(score, 3), "level": risk_level(score)})
            peak_risk = max(point["risk_score"] for point in curve)
            expected_mechanism = (
                "结构面滑移扰动型"
                if scenario["scenario_key"] == "scenario_d" or vector["geology_structure_index"] > 0.55
                else "掘进卸荷-能量突释型"
                if vector["tbm_disturbance_index"] > 0.52
                else "高地应力-脆性破裂型"
            )
            branches.append(
                {
                    "scenario_key": scenario["scenario_key"],
                    "name": scenario["name"],
                    "control_action": scenario["suggested_action"],
                    "risk_scores": risk_scores,
                    "risk_curve": curve,
                    "high_risk_interval": {
                        "chainage_start": round(chainage - 35, 1),
                        "chainage_end": round(chainage + (110 if scenario["scenario_key"] == "scenario_d" else 80), 1),
                        "label": "K{:.0f}-K{:.0f}".format(
                            chainage - 35,
                            chainage + (110 if scenario["scenario_key"] == "scenario_d" else 80),
                        ),
                    },
                    "expected_dominant_mechanism": expected_mechanism,
                    "suggested_action": scenario["suggested_action"],
                    "peak_risk": round(peak_risk, 3),
                }
            )
        best = min(branches, key=lambda branch: branch["peak_risk"])
        return {
            "score": round(base_score, 3),
            "level": risk_level(base_score),
            "branches": branches,
            "experiment_scenarios": branches,
            "best_scenario": best,
            "findings": [
                f"共生成 {len(branches)} 组反事实实验分支",
                f"未来 1h/3h/6h 最低峰值风险来自：{best['name']}，峰值 {best['peak_risk']:.2f}",
            ],
        }


class WarningDecisionAgent(DomainAgent):
    def analyze(self, state: dict[str, Any]) -> dict[str, Any]:
        return self.invoke(state, self._logic)

    def _logic(self, state: dict[str, Any]) -> dict[str, Any]:
        agents = state["agent_outputs"]
        score = clamp(
            0.27 * agents["microseismic"]["score"]
            + 0.18 * agents["tbm"]["score"]
            + 0.21 * agents["geology"]["score"]
            + 0.20 * agents["mechanism"]["score"]
            + 0.14 * agents["simulation"]["best_scenario"]["peak_risk"]
        )
        chainage = state["chainage"]
        best = agents["simulation"]["best_scenario"]
        measures = [
            best["suggested_action"],
            "对 K{:.0f}-K{:.0f} 区段提高微震采样与人工巡检频次".format(chainage - 40, chainage + 90),
        ]
        if risk_level(score) in {"高", "严重"}:
            measures.append("执行短进尺、弱扰动掘进，并复核掌子面前方构造")
        confidence = clamp(0.52 + 0.14 * len(state["geology"]["active_structures"]) + 0.18 * state["data_quality"]["coverage"])
        risk_interval = {
            "chainage_start": round(chainage - 40, 1),
            "chainage_end": round(chainage + 90, 1),
            "label": "K{:.0f}-K{:.0f}".format(chainage - 40, chainage + 90),
        }
        return {
            "risk_score": round(score, 3),
            "risk_level": risk_level(score),
            "final_risk_level": risk_level(score),
            "risk_interval": risk_interval,
            "risk_position": risk_interval,
            "dominant_mechanism_path": agents["mechanism"]["dominant_path"],
            "risk_mechanism": agents["mechanism"]["dominant_mechanism"],
            "recommended_measures": measures,
            "recommended_plan": {
                "scenario": best["name"],
                "action": best["suggested_action"],
                "measures": measures,
            },
            "plan_confidence": round(confidence, 3),
            "confidence": round(confidence, 3),
            "supplemental_monitoring_suggestions": [
                "微震定位误差复核",
                "TBM 推力/扭矩高频采样",
                "掌子面前方 10-30m 超前地质预报",
            ],
            "best_scenario": best["name"],
        }


class FeedbackCorrectionAgent(DomainAgent):
    def analyze(self, state: dict[str, Any]) -> dict[str, Any]:
        return self.invoke(state, self._logic)

    def _logic(self, state: dict[str, Any]) -> dict[str, Any]:
        vector = state["state_vector"]
        follow_up = []
        if state["microseismic"]["recent_event_count"] < 12:
            follow_up.append("补采高频微震事件，重点覆盖掌子面前方 150 m")
        if vector["geology_structure_index"] > 0.45:
            follow_up.append("补充超前地质预报或钻孔成像，校正结构面连续性")
        if vector["tbm_disturbance_index"] > 0.45:
            follow_up.append("同步采集刀盘振动、推进油缸压力和出渣粒径")
        if not follow_up:
            follow_up.append("保持微震、TBM 与地质编录三源同步更新")
        return {
            "parameter_updates": {
                "mechanism_weight_delta": round(0.08 * vector["geology_structure_index"], 3),
                "tbm_disturbance_weight_delta": round(0.06 * vector["tbm_disturbance_index"], 3),
                "microseismic_energy_weight_delta": round(0.07 * vector["microseismic_energy_acceleration"], 3),
            },
            "data_to_collect_next": follow_up,
            "findings": ["根据后续真实监测结果回写 agent 权重和机理路径置信度"],
        }


class RockburstLabOrchestrator:
    def __init__(
        self,
        default_microseismic_path: Path,
        default_tbm_path: Path,
        default_geology_path: Path,
        runtime: OpenClawRuntime | None = None,
    ) -> None:
        self.default_microseismic_path = default_microseismic_path
        self.default_tbm_path = default_tbm_path
        self.default_geology_path = default_geology_path
        self.runtime = runtime or OpenClawRuntime()
        self.microseismic_agent = MicroseismicSensingAgent("微震感知智能体", "微震时空聚集与能量演化分析", self.runtime)
        self.tbm_agent = TbmConditionAgent("掘进工况智能体", "TBM 参数异常与扰动强度分析", self.runtime)
        self.geology_agent = GeologyCognitionAgent("地质认知智能体", "断层、蚀变带和结构面风险解释", self.runtime)
        self.mechanism_agent = MechanismMatchingAgent("机理匹配智能体", "当前状态与机理图谱节点路径匹配", self.runtime)
        self.simulation_agent = SimulationExperimentAgent("推演实验智能体", "数字实验室反事实情景推演", self.runtime)
        self.decision_agent = WarningDecisionAgent("预警决策智能体", "风险等级、危险区段和处置方案输出", self.runtime)
        self.feedback_agent = FeedbackCorrectionAgent("反馈校正智能体", "监测反馈下的参数和机理权重修正", self.runtime)

    def run(
        self,
        microseismic_path: str | Path | None = None,
        tbm_path: str | Path | None = None,
        geology_path: str | Path | None = None,
    ) -> dict[str, Any]:
        micro_path = Path(microseismic_path) if microseismic_path else self.default_microseismic_path
        tbm_path_obj = Path(tbm_path) if tbm_path else self.default_tbm_path
        geology_path_obj = Path(geology_path) if geology_path else self.default_geology_path

        microseismic = self._load_microseismic(micro_path)
        tbm = self._load_tbm(tbm_path_obj)
        geology = self._load_geology(geology_path_obj)
        state = self._build_state(microseismic, tbm, geology)

        micro_result = self.microseismic_agent.analyze(state)
        tbm_result = self.tbm_agent.analyze(state)
        geology_result = self.geology_agent.analyze(state)
        mechanism_result = self.mechanism_agent.analyze(state)
        simulation_result = self.simulation_agent.analyze(state)

        state["agent_outputs"] = {
            "microseismic": micro_result,
            "tbm": tbm_result,
            "geology": geology_result,
            "mechanism": mechanism_result,
            "simulation": simulation_result,
        }
        decision_result = self.decision_agent.analyze(state)
        feedback_result = self.feedback_agent.analyze(state)

        return {
            "lab_name": "rockburst-agent-lab",
            "generated_at": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
            "inputs": {
                "microseismic": str(micro_path),
                "tbm": str(tbm_path_obj),
                "geology": str(geology_path_obj),
            },
            "state_snapshot": state["state_snapshot"],
            "state": state,
            "agents": state["agent_outputs"] | {
                "microseismic_agent": micro_result,
                "tbm_agent": tbm_result,
                "geology_agent": geology_result,
                "mechanism_agent": mechanism_result,
                "experiment_agent": simulation_result,
                "warning_agent": decision_result,
                "feedback_agent": feedback_result,
                "decision": decision_result,
                "feedback": feedback_result,
            },
            "closed_loop_output": {
                **decision_result,
                "data_to_collect_next": feedback_result["data_to_collect_next"],
                "monitoring_suggestions": decision_result["supplemental_monitoring_suggestions"] + feedback_result["data_to_collect_next"],
                "parameter_updates": feedback_result["parameter_updates"],
            },
        }

    def _load_microseismic(self, path: Path) -> list[dict[str, Any]]:
        events = []
        for row in read_table(path):
            events.append(
                {
                    "timestamp": row["timestamp"],
                    "time": parse_time(row["timestamp"]),
                    "x": parse_float(row.get("x")),
                    "y": parse_float(row.get("y")),
                    "z": parse_float(row.get("z")),
                    "energy": parse_float(row.get("energy_j")),
                    "magnitude": parse_float(row.get("magnitude")),
                    "mechanism": row.get("mechanism", "unknown"),
                }
            )
        return sorted(events, key=lambda item: item["time"])

    def _load_tbm(self, path: Path) -> list[dict[str, Any]]:
        records = []
        for row in read_table(path):
            records.append(
                {
                    "timestamp": row["timestamp"],
                    "time": parse_time(row["timestamp"]),
                    "chainage": parse_float(row.get("chainage_m")),
                    "thrust": parse_float(row.get("thrust_kn")),
                    "torque": parse_float(row.get("torque_knm")),
                    "advance_rate": parse_float(row.get("advance_rate_mm_min")),
                    "rpm": parse_float(row.get("rpm")),
                    "penetration": parse_float(row.get("penetration_mm_rev")),
                    "support_pressure": parse_float(row.get("support_pressure_mpa")),
                }
            )
        return sorted(records, key=lambda item: item["time"])

    def _load_geology(self, path: Path) -> dict[str, Any]:
        with path.open("r", encoding="utf-8") as file:
            return json.load(file)

    def _build_state(
        self,
        microseismic: list[dict[str, Any]],
        tbm: list[dict[str, Any]],
        geology: dict[str, Any],
    ) -> dict[str, Any]:
        latest_tbm = tbm[-1]
        chainage = latest_tbm["chainage"]
        recent_micro = microseismic[-12:] if len(microseismic) >= 12 else microseismic
        recent_energy = [event["energy"] for event in recent_micro]
        first_half = recent_energy[: max(1, len(recent_energy) // 2)]
        second_half = recent_energy[max(1, len(recent_energy) // 2) :]
        energy_sum = sum(recent_energy)
        event_count = len(recent_micro)
        max_energy = max(recent_energy) if recent_energy else 0.0
        centroid = {
            "x": round(mean(event["x"] for event in recent_micro), 3),
            "y": round(mean(event["y"] for event in recent_micro), 3),
            "z": round(mean(event["z"] for event in recent_micro), 3),
        }
        front_events = sum(1 for event in recent_micro if event["x"] >= chainage)
        left_events = sum(1 for event in recent_micro if event["y"] < 0)
        top_events = sum(1 for event in recent_micro if event["z"] >= centroid["z"])
        mechanism_counts = Counter(event["mechanism"] for event in recent_micro)
        acceleration = normalize(mean(second_half) - mean(first_half), 0, 18000)

        thrust_values = [record["thrust"] for record in tbm]
        torque_values = [record["torque"] for record in tbm]
        latest_thrust_anomaly = self._z_anomaly(latest_tbm["thrust"], thrust_values)
        latest_torque_anomaly = self._z_anomaly(latest_tbm["torque"], torque_values)
        advance_pressure_index = clamp(
            0.58 * normalize(latest_tbm["advance_rate"], 35, 72)
            + 0.42 * normalize(latest_tbm["support_pressure"], 1.8, 3.2)
        )
        disturbance = clamp(0.38 * latest_thrust_anomaly + 0.34 * latest_torque_anomaly + 0.28 * advance_pressure_index)
        tbm_window = tbm[-8:] if len(tbm) >= 8 else tbm
        fluctuation_amplitude = {
            "advance_rate": round(relative_range(record["advance_rate"] for record in tbm_window), 3),
            "rpm": round(relative_range(record["rpm"] for record in tbm_window), 3),
            "thrust": round(relative_range(record["thrust"] for record in tbm_window), 3),
            "torque": round(relative_range(record["torque"] for record in tbm_window), 3),
            "penetration": round(relative_range(record["penetration"] for record in tbm_window), 3),
        }
        anomaly_score = clamp(
            0.22 * normalize(z_score(latest_tbm["advance_rate"], [record["advance_rate"] for record in tbm]), 0, 3)
            + 0.18 * normalize(z_score(latest_tbm["rpm"], [record["rpm"] for record in tbm]), 0, 3)
            + 0.25 * latest_thrust_anomaly
            + 0.25 * latest_torque_anomaly
            + 0.10 * normalize(z_score(latest_tbm["penetration"], [record["penetration"] for record in tbm]), 0, 3)
        )

        stress_mpa = parse_float(geology.get("in_situ_stress", {}).get("max_principal_mpa"))
        brittleness = parse_float(geology.get("rock_mass", {}).get("brittleness_index"))
        active_structures = self._active_structures(geology, chainage)
        structure_index = clamp(
            sum(parse_float(item.get("risk_weight"), 0.35) * parse_float(item.get("confidence"), 0.75) for item in active_structures)
        )
        active_types = {item["type"] for item in active_structures}
        current_body = geology.get("rock_mass", {}).get("lithology", "unknown")
        structural_attitude_encoding = [
            {
                "type": item["type"],
                "dip_degree": item["dip"],
                "dip_sin": round(math.sin(math.radians(item["dip"])), 3),
                "dip_cos": round(math.cos(math.radians(item["dip"])), 3),
                "confidence": item["confidence"],
            }
            for item in active_structures
        ]
        geologic_complexity_score = clamp(
            0.42 * structure_index
            + 0.28 * normalize(stress_mpa, 35, 70)
            + 0.20 * normalize(brittleness, 0.45, 0.90)
            + 0.10 * min(1.0, len(active_structures) / 3)
        )
        structural_risk_tags = []
        if "fault" in active_types:
            structural_risk_tags.append("靠近断层")
        if "alteration_zone" in active_types:
            structural_risk_tags.append("靠近蚀变带")
        if "joint_set" in active_types or "structural_plane" in active_types:
            structural_risk_tags.append("靠近结构面")
        if not structural_risk_tags:
            structural_risk_tags.append("未命中显著构造")
        vector = {
            "microseismic_event_density": normalize(len(recent_micro), 2, 14),
            "microseismic_energy_index": normalize(math.log10(energy_sum + 1), 4.4, 6.4),
            "microseismic_cluster_index": self._micro_cluster_index(recent_micro),
            "microseismic_energy_acceleration": acceleration,
            "microseismic_mechanism_index": self._micro_mechanism_index(recent_micro),
            "tbm_disturbance_index": disturbance,
            "tbm_thrust_anomaly": latest_thrust_anomaly,
            "tbm_torque_anomaly": latest_torque_anomaly,
            "tbm_advance_pressure_index": advance_pressure_index,
            "geology_structure_index": structure_index,
            "geology_stress_index": normalize(stress_mpa, 35, 70),
            "geology_brittleness_index": normalize(brittleness, 0.45, 0.90),
        }
        time_window = {
            "start": recent_micro[0]["timestamp"] if recent_micro else None,
            "end": recent_micro[-1]["timestamp"] if recent_micro else None,
            "duration_min": round((recent_micro[-1]["time"] - recent_micro[0]["time"]).total_seconds() / 60, 1) if len(recent_micro) > 1 else 0.0,
        }
        microseismic_features = {
            "time_window": time_window,
            "event_count": event_count,
            "cumulative_energy_j": round(energy_sum, 3),
            "max_energy_j": round(max_energy, 3),
            "spatial_centroid": centroid,
            "front_event_ratio": round(safe_ratio(front_events, event_count), 3),
            "behind_event_ratio": round(1 - safe_ratio(front_events, event_count), 3) if event_count else 0.0,
            "distribution_ratios": {
                "left": round(safe_ratio(left_events, event_count), 3),
                "right": round(1 - safe_ratio(left_events, event_count), 3) if event_count else 0.0,
                "arch_top": round(safe_ratio(top_events, event_count), 3),
                "arch_bottom": round(1 - safe_ratio(top_events, event_count), 3) if event_count else 0.0,
            },
            "mechanism_ratios": {
                "shear": round(safe_ratio(mechanism_counts.get("shear", 0), event_count), 3),
                "tensile": round(safe_ratio(mechanism_counts.get("tensile", 0), event_count), 3),
                "mixed": round(safe_ratio(mechanism_counts.get("mixed", 0), event_count), 3),
                "unknown": round(safe_ratio(mechanism_counts.get("unknown", 0), event_count), 3),
            },
            "event_density_index": round(vector["microseismic_event_density"], 3),
            "energy_index": round(vector["microseismic_energy_index"], 3),
            "energy_acceleration_index": round(acceleration, 3),
        }
        tbm_features = {
            "advance_rate": latest_tbm["advance_rate"],
            "cutterhead_rpm": latest_tbm["rpm"],
            "thrust": latest_tbm["thrust"],
            "torque": latest_tbm["torque"],
            "penetration": latest_tbm["penetration"],
            "fluctuation_amplitude": fluctuation_amplitude,
            "anomaly_score": round(anomaly_score, 3),
            "disturbance_index": round(disturbance, 3),
        }
        geology_features = {
            "current_geologic_body_type": current_body,
            "near_fault": "fault" in active_types,
            "near_alteration_zone": "alteration_zone" in active_types,
            "near_structural_plane": bool({"joint_set", "structural_plane"} & active_types),
            "structural_attitude_encoding": structural_attitude_encoding,
            "geologic_complexity_score": round(geologic_complexity_score, 3),
            "structural_risk_tags": structural_risk_tags,
            "active_structures": active_structures,
            "all_structures": geology.get("structures", []),
        }
        data_coverage = clamp(
            0.34 * min(1.0, len(microseismic) / 18)
            + 0.33 * min(1.0, len(tbm) / 14)
            + 0.33 * (1.0 if geology else 0.0)
        )
        risk_context = {
            "current_time_window": time_window,
            "face_chainage": chainage,
            "data_coverage": round(data_coverage, 3),
            "microseismic_activity_prior": round(
                clamp(0.42 * vector["microseismic_event_density"] + 0.38 * vector["microseismic_energy_index"] + 0.20 * acceleration),
                3,
            ),
            "tbm_disturbance_prior": round(disturbance, 3),
            "geology_modifier": round(0.85 + 0.42 * geologic_complexity_score, 3),
            "composite_context_score": round(
                clamp(0.38 * vector["microseismic_energy_index"] + 0.28 * disturbance + 0.24 * geologic_complexity_score + 0.10 * vector["microseismic_cluster_index"]),
                3,
            ),
        }
        state_snapshot = {
            "timestamp": latest_tbm["timestamp"],
            "chainage": chainage,
            "microseismic_features": microseismic_features,
            "tbm_features": tbm_features,
            "geology_features": geology_features,
            "risk_context": risk_context,
        }
        return {
            "timestamp": latest_tbm["timestamp"],
            "chainage": chainage,
            "state_snapshot": state_snapshot,
            "state_vector": {key: round(value, 4) for key, value in vector.items()},
            "state_vector_ordered": [{"name": name, "value": round(value, 3)} for name, value in vector.items()],
            "microseismic": {
                "recent_event_count": event_count,
                "recent_energy_sum": energy_sum,
                "events": [{key: value for key, value in event.items() if key != "time"} for event in microseismic],
            },
            "tbm": {
                "latest": {key: value for key, value in latest_tbm.items() if key != "time"},
                "series": [{key: value for key, value in record.items() if key != "time"} for record in tbm],
            },
            "geology": {
                "in_situ_stress_mpa": stress_mpa,
                "active_structures": active_structures,
                "all_structures": geology.get("structures", []),
                "mechanism_graph": geology.get("mechanism_graph", {}),
            },
            "data_quality": {
                "coverage": round(data_coverage, 3),
            },
        }

    def _active_structures(self, geology: dict[str, Any], chainage: float) -> list[dict[str, Any]]:
        active = []
        for item in geology.get("structures", []):
            start = parse_float(item.get("chainage_start"))
            end = parse_float(item.get("chainage_end"))
            influence = parse_float(item.get("influence_radius_m"), 40)
            if start - influence <= chainage <= end + influence:
                active.append(
                    {
                        "type": item.get("type", "structure"),
                        "chainage_start": start,
                        "chainage_end": end,
                        "dip": parse_float(item.get("dip")),
                        "confidence": parse_float(item.get("confidence"), 0.7),
                        "risk_weight": parse_float(item.get("risk_weight"), 0.35),
                        "influence_radius_m": influence,
                        "description": item.get("description", ""),
                    }
                )
        return active

    def _micro_cluster_index(self, events: list[dict[str, Any]]) -> float:
        if len(events) < 2:
            return 0.0
        distances = []
        for idx, event in enumerate(events):
            for other in events[idx + 1 :]:
                distances.append(
                    math.sqrt(
                        (event["x"] - other["x"]) ** 2
                        + (event["y"] - other["y"]) ** 2
                        + (event["z"] - other["z"]) ** 2
                    )
                )
        return clamp(1 - mean(distances) / 130)

    def _micro_mechanism_index(self, events: list[dict[str, Any]]) -> float:
        if not events:
            return 0.0
        risky = sum(1 for event in events if event["mechanism"] in {"shear", "tensile", "mixed"})
        mixed_bonus = 0.12 if any(event["mechanism"] == "mixed" for event in events) else 0.0
        return clamp(risky / len(events) + mixed_bonus)

    def _z_anomaly(self, latest: float, values: list[float]) -> float:
        sigma = stdev(values)
        if sigma <= 1e-9:
            return 0.0
        return clamp(abs(latest - mean(values)) / (3.0 * sigma))
